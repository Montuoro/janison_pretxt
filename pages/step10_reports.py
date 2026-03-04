"""Step 10: Reports — Wright map, fit, Person-Item distribution, Guttman."""

from dash import html, dcc, callback, Input, Output, State, no_update
import dash_bootstrap_components as dbc
from dash import dash_table
import pandas as pd
import numpy as np
import json
import plotly.graph_objects as go
from plotly.subplots import make_subplots

from core.report_builder import build_item_params_table, identify_misfit_items
from components.page_size_selector import page_size_selector, register_page_size_callback

# Pre-register page-size callbacks for dynamically created tables
register_page_size_callback("report-params-page-size", "report-params-table")
# Guttman table uses virtualization (no pagination), so no page-size callback needed

layout = dbc.Container([
    html.H3("Step 10: Reports", className="mb-3"),

    dbc.Tabs([
        dbc.Tab(label="Item Parameters", tab_id="tab-params"),
        dbc.Tab(label="Wright Map", tab_id="tab-wright"),
        dbc.Tab(label="Test Statistics", tab_id="tab-stats"),
        dbc.Tab(label="Person-Item Distribution", tab_id="tab-person-item"),
        dbc.Tab(label="Guttman Pattern", tab_id="tab-guttman"),
        dbc.Tab(label="Export", tab_id="tab-export"),
    ], id="report-tabs", active_tab="tab-params"),

    html.Div([
        html.Div(id="report-tab-content", className="report-tab-content"),

        # Page-size selectors (always in DOM; visibility toggled by tab callback)
        html.Div(id="report-params-page-size-container", children=[
            page_size_selector("report-params-page-size", default=15),
        ], style={"display": "none"}),
        html.Div(id="report-guttman-page-size-container", children=[
            page_size_selector("report-guttman-page-size", default=50),
        ], style={"display": "none"}),
    ], className="table-with-pager"),
], fluid=True)


@callback(
    Output("report-tab-content", "children"),
    Output("store-completed-steps", "data", allow_duplicate=True),
    Output("report-params-page-size-container", "style"),
    Output("report-guttman-page-size-container", "style"),
    Input("report-tabs", "active_tab"),
    State("store-btl-results", "data"),
    State("store-tam-results", "data"),
    State("store-abilities", "data"),
    State("store-response-matrix", "data"),
    State("store-discrimination", "data"),
    State("store-completed-steps", "data"),
    prevent_initial_call=True,
)
def render_report_tab(tab, btl_json, tam_json, abilities_json, resp_json, discrim_json, completed):
    completed = completed or []
    if 10 not in completed:
        completed = sorted(set(completed + [10]))

    hide = {"display": "none"}
    show = {"display": "block"}
    params_vis = show if tab == "tab-params" else hide
    guttman_vis = hide  # Guttman uses virtualization, no page-size selector

    if not btl_json:
        return dbc.Alert("Complete earlier steps first.", color="warning"), completed, hide, hide

    items_df = pd.read_json(btl_json, orient="split")
    tam_results = json.loads(tam_json) if tam_json else None
    abilities = np.array(json.loads(abilities_json)) if abilities_json else None
    resp_df = pd.read_json(resp_json, orient="split") if resp_json else None

    if discrim_json:
        from core.discrimination import apply_discrimination
        flags = json.loads(discrim_json)
        items_df = apply_discrimination(items_df, flags)

    if tab == "tab-params":
        return _render_params(items_df, tam_results), completed, params_vis, guttman_vis
    elif tab == "tab-wright":
        return _render_wright_controls(items_df, tam_results, abilities), completed, params_vis, guttman_vis
    elif tab == "tab-stats":
        return _render_stats(tam_results), completed, params_vis, guttman_vis
    elif tab == "tab-person-item":
        return _render_person_item_controls(abilities, items_df, tam_results), completed, params_vis, guttman_vis
    elif tab == "tab-guttman":
        return _render_guttman(resp_df, abilities, items_df), completed, params_vis, guttman_vis
    elif tab == "tab-export":
        return _render_export(items_df, tam_results), completed, params_vis, guttman_vis

    return html.Div(), completed


def _render_params(items_df, tam_results):
    params_df = build_item_params_table(items_df, tam_results)
    params_df = identify_misfit_items(params_df)

    # Color-code misfit
    style_conditions = []
    if "infit" in params_df.columns:
        style_conditions.extend([
            {"if": {"filter_query": "{infit} > 1.3", "column_id": "infit"},
             "backgroundColor": "#f8d7da"},
            {"if": {"filter_query": "{infit} < 0.7", "column_id": "infit"},
             "backgroundColor": "#f8d7da"},
            {"if": {"filter_query": "{outfit} > 1.3", "column_id": "outfit"},
             "backgroundColor": "#f8d7da"},
            {"if": {"filter_query": "{outfit} < 0.7", "column_id": "outfit"},
             "backgroundColor": "#f8d7da"},
        ])

    # Round numeric columns
    for col in ["btl_difficulty", "btl_se", "rasch_difficulty", "rasch_se", "infit", "outfit",
                "infit_t", "outfit_t", "discrimination"]:
        if col in params_df.columns:
            params_df[col] = pd.to_numeric(params_df[col], errors="coerce").round(3)

    display_cols = [c for c in params_df.columns if c not in ("misfit",)]

    return html.Div([
        html.H5("Item Parameters"),
        dash_table.DataTable(
            id="report-params-table",
            columns=[{"name": c.replace("_", " ").title(), "id": c} for c in display_cols],
            data=params_df[display_cols].to_dict("records"),
            style_table={"overflowX": "auto"},
            style_cell={"textAlign": "center", "padding": "6px", "fontSize": "0.8rem"},
            style_cell_conditional=[
                {"if": {"column_id": "item_id"}, "textAlign": "left"},
                {"if": {"column_id": "stem"}, "textAlign": "left", "maxWidth": "250px",
                 "whiteSpace": "normal"},
            ],
            style_header={"backgroundColor": "#e9ecef", "fontWeight": "600",
                           "textAlign": "center"},
            style_data_conditional=style_conditions,
            page_action="none",
            sort_action="native",
            filter_action="native",
        ),
    ])


_WRIGHT_INTERVAL_OPTIONS = [
    {"label": f"{v:.1f} logits", "value": v}
    for v in [0.4, 0.6, 0.8, 1.0, 1.2, 1.5, 2.0]
]


def _render_wright_controls(items_df, tam_results, abilities):
    """Return interval selector + initial Wright map."""
    if abilities is None or len(abilities) == 0:
        return dbc.Alert("Generate person data first (Step 8).", color="warning")

    # Get item data
    if tam_results and "item_params" in tam_results:
        item_diffs = np.array([p["xsi"] for p in tam_results["item_params"]])
        item_ids = [p.get("item", items_df["item_id"].values[i])
                    for i, p in enumerate(tam_results["item_params"])]
    else:
        item_diffs = items_df["difficulty"].values
        item_ids = items_df["item_id"].values.tolist()

    # Build item_id → stem lookup for tooltips
    stem_lookup = dict(zip(items_df["item_id"].astype(str), items_df["stem"].astype(str)))

    initial_fig = _build_wright_figure(abilities, item_diffs, item_ids, stem_lookup, 0.8)

    return html.Div([
        html.H5("Wright Map"),
        dbc.Row([
            dbc.Col([
                html.Small("Interval:", className="text-muted me-2"),
                dcc.Dropdown(
                    id="wright-interval",
                    options=_WRIGHT_INTERVAL_OPTIONS,
                    value=0.8,
                    clearable=False,
                    searchable=False,
                    style={"width": "140px", "display": "inline-block"},
                ),
            ], width="auto", className="d-flex align-items-center"),
        ], className="mb-2"),
        dcc.Graph(id="wright-map-chart", figure=initial_fig),
    ])


def _build_wright_figure(abilities, item_diffs, item_ids, stem_lookup, bin_width):
    """Build RUMM2030-style Wright map figure."""
    all_vals = np.concatenate([abilities, item_diffs])
    lo = np.floor(all_vals.min()) - bin_width * 0.5
    hi = np.ceil(all_vals.max()) + bin_width * 0.5
    bin_edges = np.arange(lo, hi + bin_width, bin_width)
    bin_centres = (bin_edges[:-1] + bin_edges[1:]) / 2

    fig = make_subplots(
        rows=1, cols=2,
        column_widths=[0.30, 0.70],
        shared_yaxes=True,
        horizontal_spacing=0.02,
    )

    # ── Left panel: Person histogram ──────────────────────────────────────
    person_counts, _ = np.histogram(abilities, bins=bin_edges)
    fig.add_trace(
        go.Bar(
            y=bin_centres,
            x=person_counts,
            orientation="h",
            marker_color="rgba(100, 100, 180, 0.5)",
            marker_line=dict(color="#333366", width=1),
            marker_pattern=dict(shape="/", fgcolor="#333366", bgcolor="rgba(100,100,180,0.3)"),
            name="Persons",
            text=[str(c) if c > 0 else "" for c in person_counts],
            textposition="outside",
            textfont=dict(size=10),
            hovertemplate="Location: %{y:.2f}<br>Count: %{x}<extra></extra>",
            width=bin_width * 0.85,
        ),
        row=1, col=1,
    )
    max_count = int(person_counts.max()) + 5
    fig.update_xaxes(
        autorange="reversed", range=[max_count, 0],
        showticklabels=False, showgrid=False,
        row=1, col=1,
    )

    # ── Right panel: Item IDs as scatter text with hover stems ────────────
    item_bin_idx = np.digitize(item_diffs, bin_edges) - 1
    item_bin_idx = np.clip(item_bin_idx, 0, len(bin_centres) - 1)

    # Group items by bin
    bin_items = {i: [] for i in range(len(bin_centres))}
    for idx, b in enumerate(item_bin_idx):
        bin_items[b].append((item_ids[idx], item_diffs[idx]))

    # Place items as scatter points with text labels and hover stems
    item_x = []
    item_y = []
    item_text = []
    item_hover = []
    for b, items in bin_items.items():
        if not items:
            continue
        for j, (iid, diff) in enumerate(items):
            item_x.append(j)
            item_y.append(bin_centres[b])
            item_text.append(str(iid))
            stem = stem_lookup.get(str(iid), "")
            if len(stem) > 80:
                stem = stem[:80] + "..."
            item_hover.append(f"<b>{iid}</b><br>Difficulty: {diff:.2f}<br>{stem}")

    fig.add_trace(
        go.Scatter(
            x=item_x,
            y=item_y,
            mode="text",
            text=item_text,
            textposition="middle right",
            textfont=dict(size=13, family="Courier New, monospace"),
            hovertext=item_hover,
            hoverinfo="text",
            name="Items",
        ),
        row=1, col=2,
    )

    max_in_bin = max((len(ids) for ids in bin_items.values()), default=1)
    fig.update_xaxes(
        range=[-0.5, max_in_bin + 0.5],
        showticklabels=False, showgrid=False,
        row=1, col=2,
    )

    # ── Shared y-axis ─────────────────────────────────────────────────────
    fig.update_yaxes(
        title_text="Location (logits)",
        dtick=bin_width,
        showgrid=False,
        row=1, col=1,
    )
    fig.update_yaxes(showgrid=False, row=1, col=2)

    # Column headers
    fig.add_annotation(
        x=0.12, y=1.04, xref="paper", yref="paper",
        text="<b>Persons</b>", showarrow=False, font=dict(size=12),
    )
    fig.add_annotation(
        x=0.68, y=1.04, xref="paper", yref="paper",
        text="<b>Items</b>", showarrow=False, font=dict(size=12),
    )

    # Central vertical line
    fig.add_shape(
        type="line",
        x0=0.30, x1=0.30, y0=0, y1=1,
        xref="paper", yref="paper",
        line=dict(color="#333366", width=2),
    )

    n_bins = len(bin_centres)
    chart_height = max(550, n_bins * 45)

    fig.update_layout(
        template="plotly_white",
        height=chart_height,
        showlegend=False,
        margin=dict(t=50, b=30, l=50, r=20),
    )

    return fig


@callback(
    Output("wright-map-chart", "figure"),
    Input("wright-interval", "value"),
    State("store-btl-results", "data"),
    State("store-tam-results", "data"),
    State("store-abilities", "data"),
    State("store-items", "data"),
    prevent_initial_call=True,
)
def update_wright_map(bin_width, btl_json, tam_json, abilities_json, items_json):
    if not btl_json or not abilities_json:
        return go.Figure()

    items_df = pd.read_json(btl_json, orient="split")
    tam_results = json.loads(tam_json) if tam_json else None
    abilities = np.array(json.loads(abilities_json))
    bin_width = float(bin_width or 0.8)

    if tam_results and "item_params" in tam_results:
        item_diffs = np.array([p["xsi"] for p in tam_results["item_params"]])
        item_ids = [p.get("item", items_df["item_id"].values[i])
                    for i, p in enumerate(tam_results["item_params"])]
    else:
        item_diffs = items_df["difficulty"].values
        item_ids = items_df["item_id"].values.tolist()

    # Build stem lookup — use original items store if available
    if items_json:
        orig_df = pd.read_json(items_json, orient="split")
        stem_lookup = dict(zip(orig_df["item_id"].astype(str), orig_df["stem"].astype(str)))
    else:
        stem_lookup = dict(zip(items_df["item_id"].astype(str),
                               items_df["stem"].astype(str) if "stem" in items_df.columns else [""] * len(items_df)))

    return _build_wright_figure(abilities, item_diffs, item_ids, stem_lookup, bin_width)


def _render_stats(tam_results):
    if not tam_results:
        return dbc.Alert("Run Rasch analysis in Step 9 first.", color="warning")

    rel = tam_results.get("reliability", {})

    return html.Div([
        html.H5("Test Statistics"),
        dbc.Row([
            dbc.Col(dbc.Card(dbc.CardBody([
                html.Div(f"{rel.get('cronbach_alpha', 0):.3f}", className="stat-value"),
                html.Div("Cronbach's Alpha", className="stat-label"),
            ]), className="stat-card"), md=4),
            dbc.Col(dbc.Card(dbc.CardBody([
                html.Div(f"{rel.get('wle_rel', 0):.3f}", className="stat-value"),
                html.Div("WLE PSI Reliability", className="stat-label"),
            ]), className="stat-card"), md=4),
            dbc.Col(dbc.Card(dbc.CardBody([
                html.Div(f"{rel.get('eap_rel', 0):.3f}", className="stat-value"),
                html.Div("EAP Reliability", className="stat-label"),
            ]), className="stat-card"), md=4),
        ]),
        html.Hr(),
        dbc.Row([
            dbc.Col(dbc.Card(dbc.CardBody([
                html.Div(f"{tam_results.get('n_items', 0)}", className="stat-value"),
                html.Div("Items", className="stat-label"),
            ]), className="stat-card"), md=6),
            dbc.Col(dbc.Card(dbc.CardBody([
                html.Div(f"{tam_results.get('n_persons', 0)}", className="stat-value"),
                html.Div("Persons", className="stat-label"),
            ]), className="stat-card"), md=6),
        ]),
    ])


# ── Person-Item Threshold Distribution (replaces Gap Analysis) ───────────────

_INTERVAL_OPTIONS = [
    {"label": f"{v:.1f} logits", "value": v}
    for v in [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8]
]


def _render_person_item_controls(abilities, items_df, tam_results):
    """Return the interval selector + initial chart."""
    if abilities is None or len(abilities) == 0:
        return dbc.Alert("Generate person data first (Step 8).", color="warning")

    # Build initial figure so chart isn't blank on load
    if tam_results and "item_params" in tam_results:
        item_diffs = np.array([p["xsi"] for p in tam_results["item_params"]])
    else:
        item_diffs = items_df["difficulty"].values
    initial_fig = _build_person_item_figure(abilities, item_diffs, 0.50)

    return html.Div([
        html.H5("Person-Item Threshold Distribution"),
        html.P("Persons on top, items on bottom — both on the same logit scale.",
               className="text-muted small"),
        dbc.Row([
            dbc.Col([
                html.Small("Interval Length:", className="text-muted me-2"),
                dcc.Dropdown(
                    id="person-item-interval",
                    options=_INTERVAL_OPTIONS,
                    value=0.50,
                    clearable=False,
                    searchable=False,
                    style={"width": "140px", "display": "inline-block"},
                ),
            ], width="auto", className="d-flex align-items-center"),
        ], className="mb-2"),
        dcc.Graph(id="person-item-chart", figure=initial_fig),
    ])


@callback(
    Output("person-item-chart", "figure"),
    Input("person-item-interval", "value"),
    State("store-btl-results", "data"),
    State("store-tam-results", "data"),
    State("store-abilities", "data"),
    prevent_initial_call=True,
)
def update_person_item_chart(bin_width, btl_json, tam_json, abilities_json):
    if not btl_json or not abilities_json:
        return go.Figure()

    items_df = pd.read_json(btl_json, orient="split")
    tam_results = json.loads(tam_json) if tam_json else None
    abilities = np.array(json.loads(abilities_json))
    bin_width = float(bin_width or 0.5)

    if tam_results and "item_params" in tam_results:
        item_diffs = np.array([p["xsi"] for p in tam_results["item_params"]])
    else:
        item_diffs = items_df["difficulty"].values

    return _build_person_item_figure(abilities, item_diffs, bin_width)


def _build_person_item_figure(abilities, item_diffs, bin_width):
    all_vals = np.concatenate([abilities, item_diffs])
    lo = np.floor(all_vals.min()) - 0.5
    hi = np.ceil(all_vals.max()) + 0.5
    bin_edges = np.arange(lo, hi + bin_width, bin_width)
    bin_centres = (bin_edges[:-1] + bin_edges[1:]) / 2

    person_counts, _ = np.histogram(abilities, bins=bin_edges)
    item_counts, _ = np.histogram(item_diffs, bins=bin_edges)

    n_persons = len(abilities)
    n_items = len(item_diffs)
    n_groups = len(bin_centres)
    person_pct = person_counts / n_persons * 100 if n_persons > 0 else person_counts
    item_pct = item_counts / n_items * 100 if n_items > 0 else item_counts

    fig = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True,
        vertical_spacing=0.03,
        row_heights=[0.65, 0.35],
    )

    fig.add_trace(
        go.Bar(
            x=bin_centres, y=person_counts,
            marker_color="#198754", opacity=0.8,
            name="Persons",
            hovertemplate="Location: %{x:.2f}<br>Count: %{y}<br>Pct: %{customdata:.1f}%",
            customdata=person_pct,
        ),
        row=1, col=1,
    )

    fig.add_annotation(
        x=0.02, y=0.98, xref="paper", yref="paper",
        text=(f"<b>PERSONS</b>  N={n_persons}  "
              f"Mean={abilities.mean():.3f}  SD={abilities.std():.3f}  "
              f"(Interval={bin_width:.2f}, {n_groups} groups)"),
        showarrow=False, font=dict(size=11),
        xanchor="left", yanchor="top",
    )

    person_max = int(person_counts.max()) + 2
    fig.update_yaxes(title_text="Frequency", range=[0, person_max], row=1, col=1)

    fig.add_trace(
        go.Bar(
            x=bin_centres, y=item_counts,
            marker_color="#0d6efd", opacity=0.8,
            name="Items",
            hovertemplate="Location: %{x:.2f}<br>Count: %{y}<br>Pct: %{customdata:.1f}%",
            customdata=item_pct,
        ),
        row=2, col=1,
    )

    item_max = int(item_counts.max()) + 2
    fig.update_yaxes(
        title_text="Frequency", autorange="reversed",
        range=[0, item_max], row=2, col=1,
    )

    fig.add_annotation(
        x=0.02, y=0.30, xref="paper", yref="paper",
        text=f"<b>ITEMS</b>  N={n_items}",
        showarrow=False, font=dict(size=11),
        xanchor="left", yanchor="top",
    )

    fig.update_xaxes(title_text="Location (logits)", row=2, col=1)

    fig.update_layout(
        title="Person-Item Threshold Distribution",
        template="plotly_white",
        height=550,
        showlegend=True,
        bargap=0.05,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )

    return fig


# ── Guttman Pattern (RUMM2030-style colored grid) ───────────────────────────

def _render_guttman(resp_df, abilities, items_df):
    """RUMM2030-style Guttman distribution: colored grid of all responses.

    Rows = persons sorted by ability (lowest first).
    Columns = items sorted by difficulty (easiest first).
    1 (correct) = blue, 0 (incorrect) = amber/orange.
    """
    if resp_df is None or abilities is None:
        return dbc.Alert("Generate response data first (Step 8).", color="warning")

    difficulties = items_df["difficulty"].values
    item_ids = items_df["item_id"].values

    # Sort items by difficulty (easiest first)
    item_order = np.argsort(difficulties)
    sorted_item_ids = item_ids[item_order]
    sorted_diffs = difficulties[item_order]
    sorted_responses = resp_df.iloc[:, item_order].values

    # Sort persons by ability (lowest first)
    person_order = np.argsort(abilities)
    sorted_abilities = abilities[person_order]
    sorted_responses = sorted_responses[person_order]

    # Build column headers: difficulty rounded
    item_col_ids = [f"item_{i}" for i in range(len(sorted_diffs))]
    item_col_names = [f"{d:.1f}" for d in sorted_diffs]

    # Build table data
    records = []
    for i in range(len(sorted_abilities)):
        row = {
            "serial": int(person_order[i]),
            "location": round(float(sorted_abilities[i]), 3),
        }
        for j, col_id in enumerate(item_col_ids):
            row[col_id] = int(sorted_responses[i, j])
        records.append(row)

    # Columns definition
    columns = [
        {"name": "Serial", "id": "serial"},
        {"name": "Location", "id": "location"},
    ] + [
        {"name": name, "id": col_id}
        for name, col_id in zip(item_col_names, item_col_ids)
    ]

    # Conditional styling: 1 = blue, 0 = amber for each item column
    style_conditions = []
    for col_id in item_col_ids:
        style_conditions.append({
            "if": {"filter_query": f"{{{col_id}}} = 1", "column_id": col_id},
            "backgroundColor": "#0d6efd", "color": "white",
        })
        style_conditions.append({
            "if": {"filter_query": f"{{{col_id}}} = 0", "column_id": col_id},
            "backgroundColor": "#e67e00", "color": "white",
        })

    # Calculate cell width to fill ~100vw
    # Reserve 50+60=110px for serial+location, 24px padding, rest for items
    n_items = len(item_col_ids)
    item_cell_w = max(28, int((1400 - 110) / n_items))

    return html.Div([
        html.Div([
            html.Span("Guttman Pattern", className="fw-bold me-3"),
            html.Span(
                f"{len(records)} persons \u00d7 {n_items} items   ",
                className="text-muted small me-2",
            ),
            html.Span("\u2588 1 correct", style={"color": "#0d6efd", "fontSize": "0.8rem", "fontWeight": "600"}),
            html.Span("  "),
            html.Span("\u2588 0 incorrect", style={"color": "#e67e00", "fontSize": "0.8rem", "fontWeight": "600"}),
            # Zoom slider
            html.Span("  Zoom:", className="text-muted ms-4 me-1",
                       style={"fontSize": "0.8rem"}),
            dcc.Slider(
                id="guttman-zoom", min=50, max=200, step=10, value=100,
                marks={50: "50%", 100: "100%", 150: "150%", 200: "200%"},
                tooltip={"placement": "bottom"},
                className="guttman-zoom-slider",
            ),
            dbc.Button("Export .xlsx", id="btn-export-guttman", color="success",
                       size="sm", className="ms-3 me-2"),
            dcc.Download(id="download-guttman-xlsx"),
            html.Button("\u2715 Close", id="btn-close-guttman",
                        className="btn btn-sm btn-outline-secondary ms-1"),
        ], className="guttman-header mb-1 d-flex align-items-center"),
        html.Div(
            dash_table.DataTable(
                id="report-guttman-table",
                columns=columns,
                data=records,
                page_action="none",
                virtualization=True,
                style_table={
                    "width": "100%",
                    "height": "calc(100vh - 50px)",
                    "overflowY": "auto",
                    "overflowX": "auto",
                },
                style_cell={
                    "textAlign": "center", "padding": "1px 2px",
                    "fontSize": "0.65rem", "fontFamily": "monospace",
                    "width": f"{item_cell_w}px",
                    "minWidth": f"{item_cell_w}px",
                    "maxWidth": f"{item_cell_w}px",
                    "lineHeight": "18px", "height": "20px",
                },
                style_cell_conditional=[
                    {"if": {"column_id": "serial"},
                     "width": "50px", "minWidth": "50px", "maxWidth": "50px",
                     "fontWeight": "600", "backgroundColor": "#f8f9fa"},
                    {"if": {"column_id": "location"},
                     "width": "60px", "minWidth": "60px", "maxWidth": "60px",
                     "fontWeight": "600", "backgroundColor": "#f8f9fa"},
                ],
                style_header={
                    "backgroundColor": "#e9ecef", "fontWeight": "600",
                    "fontSize": "0.6rem", "padding": "1px 2px",
                },
                style_data_conditional=style_conditions,
                fixed_rows={"headers": True},
            ),
            className="guttman-table-wrap",
        ),
    ], className="guttman-fullscreen")


from dash import clientside_callback

clientside_callback(
    """
    function(zoom) {
        var scale = (zoom || 100) / 100;
        var wrap = document.querySelector('.guttman-table-wrap');
        if (!wrap) return window.dash_clientside.no_update;

        var header = document.querySelector('.guttman-header');
        var headerH = header ? header.offsetHeight : 40;
        var availH = window.innerHeight - headerH - 16;
        var innerH = availH / scale;

        wrap.style.transform = 'scale(' + scale + ')';
        wrap.style.transformOrigin = 'top left';
        wrap.style.width = (100 / scale) + '%';

        // Only set height on the container chain, not on cell divs
        wrap.style.height = innerH + 'px';
        var containers = wrap.querySelectorAll(
            '.dash-spreadsheet-container, .dash-spreadsheet-inner'
        );
        containers.forEach(function(el) {
            el.style.height = innerH + 'px';
            el.style.maxHeight = 'none';
        });
        return window.dash_clientside.no_update;
    }
    """,
    Output("report-guttman-table", "id"),  # dummy output
    Input("guttman-zoom", "value"),
    prevent_initial_call=True,
)

# Force DataTable container chain to fill available height after Guttman tab renders
clientside_callback(
    """
    function(tab) {
        if (tab !== 'tab-guttman') return window.dash_clientside.no_update;
        setTimeout(function() {
            var header = document.querySelector('.guttman-header');
            var headerH = header ? header.offsetHeight : 40;
            var availH = window.innerHeight - headerH - 16;

            var wrap = document.querySelector('.guttman-table-wrap');
            if (!wrap) return;

            // Set height on wrapper and the DataTable container classes only
            wrap.style.height = availH + 'px';
            var containers = wrap.querySelectorAll(
                '.dash-spreadsheet-container, .dash-spreadsheet-inner'
            );
            containers.forEach(function(el) {
                el.style.height = availH + 'px';
                el.style.maxHeight = 'none';
            });

            // Also target the style_table wrapper (first child div of wrap)
            var tableWrap = wrap.querySelector(':scope > div');
            if (tableWrap) {
                tableWrap.style.height = availH + 'px';
            }
        }, 200);
        return window.dash_clientside.no_update;
    }
    """,
    Output("report-guttman-page-size-container", "className"),  # dummy output
    Input("report-tabs", "active_tab"),
    prevent_initial_call=True,
)


def _render_export(items_df, tam_results):
    params_df = build_item_params_table(items_df, tam_results)
    csv_string = params_df.to_csv(index=False)

    return html.Div([
        html.H5("Export"),
        dbc.Row([
            dbc.Col([
                html.H6("Item Parameters CSV"),
                html.P("Download item difficulty, fit, and discrimination parameters."),
                dbc.Button(
                    "Download CSV",
                    id="btn-download-csv",
                    color="primary",
                ),
                dcc.Download(id="download-csv"),
            ], md=6),
            dbc.Col([
                html.H6("Full Report (HTML)"),
                html.P("Coming soon: export a standalone HTML report with all charts and tables."),
                dbc.Button("Download HTML", color="secondary", disabled=True),
            ], md=6),
        ]),
        dcc.Store(id="export-csv-data", data=csv_string),
    ])


@callback(
    Output("report-tabs", "active_tab"),
    Input("btn-close-guttman", "n_clicks"),
    prevent_initial_call=True,
)
def close_guttman_fullscreen(n_clicks):
    return "tab-params"


@callback(
    Output("download-csv", "data"),
    Input("btn-download-csv", "n_clicks"),
    State("export-csv-data", "data"),
    prevent_initial_call=True,
)
def download_csv(n_clicks, csv_string):
    if not csv_string:
        return no_update
    return dict(content=csv_string, filename="pretxt_item_params.csv")


@callback(
    Output("download-guttman-xlsx", "data"),
    Input("btn-export-guttman", "n_clicks"),
    State("store-response-matrix", "data"),
    State("store-abilities", "data"),
    State("store-btl-results", "data"),
    prevent_initial_call=True,
)
def export_guttman_xlsx(n_clicks, resp_json, abilities_json, btl_json):
    """Export the Guttman pattern as a colour-coded .xlsx file."""
    if not resp_json or not abilities_json or not btl_json:
        return no_update

    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    resp_df = pd.read_json(resp_json, orient="split")
    abilities = np.array(json.loads(abilities_json))
    items_df = pd.read_json(btl_json, orient="split")

    difficulties = items_df["difficulty"].values
    item_ids = items_df["item_id"].values

    # Sort items by difficulty (easiest first)
    item_order = np.argsort(difficulties)
    sorted_item_ids = item_ids[item_order]
    sorted_diffs = difficulties[item_order]
    sorted_responses = resp_df.iloc[:, item_order].values

    # Sort persons by ability (lowest first)
    person_order = np.argsort(abilities)
    sorted_abilities = abilities[person_order]
    sorted_responses = sorted_responses[person_order]

    wb = Workbook()
    ws = wb.active
    ws.title = "Guttman Pattern"

    blue_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    amber_fill = PatternFill(start_color="E67E00", end_color="E67E00", fill_type="solid")
    white_font = Font(color="FFFFFF", size=8)
    header_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    header_font = Font(bold=True, size=8)
    center = Alignment(horizontal="center")

    # Header row
    ws.cell(row=1, column=1, value="Serial").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=1, column=2, value="Location").fill = header_fill
    ws.cell(row=1, column=2).font = header_font
    ws.cell(row=1, column=2).alignment = center
    for j, (iid, diff) in enumerate(zip(sorted_item_ids, sorted_diffs)):
        cell = ws.cell(row=1, column=j + 3, value=f"{iid}\n{diff:.1f}")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    n_persons = len(sorted_abilities)
    n_items = len(sorted_item_ids)
    for i in range(n_persons):
        ws.cell(row=i + 2, column=1, value=int(person_order[i])).alignment = center
        ws.cell(row=i + 2, column=2, value=round(float(sorted_abilities[i]), 3)).alignment = center
        for j in range(n_items):
            val = int(sorted_responses[i, j])
            cell = ws.cell(row=i + 2, column=j + 3, value=val)
            cell.alignment = center
            cell.font = white_font
            cell.fill = blue_fill if val == 1 else amber_fill

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 9
    for j in range(n_items):
        col_letter = ws.cell(row=1, column=j + 3).column_letter
        ws.column_dimensions[col_letter].width = 6

    # Freeze header row and first 2 columns
    ws.freeze_panes = "C2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    import base64
    b64 = base64.b64encode(buf.getvalue()).decode()
    return dict(
        content=b64,
        filename="guttman_pattern.xlsx",
        base64=True,
    )
